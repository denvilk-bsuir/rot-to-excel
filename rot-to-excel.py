from bs4 import BeautifulSoup
from openpyxl import load_workbook

soup = None

text = input("Enter html rot algorith file name with extension: ")
workbook = input("Enter excel filename: ")

with open(text, 'r') as f:
    soup = BeautifulSoup(f)

res = soup.find_all("table")

count = 0

wb = load_workbook(workbook)
for table in res:
    if count==0:
        count+=1
        continue
    
    ws = wb.create_sheet(f'phase_{count}')

    rows = table.find_all("tr")

    for row in rows:
        xsl_row = []
        columns = row.find_all('td')
        for column in columns:
            txt = column.get_text()
            if txt == None:
                txt = " "
            xsl_row.append(txt)
        ws.append(xsl_row)

    count+=1

wb.save(workbook)